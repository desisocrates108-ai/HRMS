from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
from database import db
from auth_utils import get_current_user, require_role, log_audit, CEO_HR_ROLES
import uuid
import io
from datetime import datetime, timezone

router = APIRouter()


# =====================================================================
# PIPELINE — Employee Database (mirrors Lead Pipeline structure)
# Stages: new -> qualified -> hr -> manager -> selected -> three_months -> joined
# Parallel: hold, rejected
# Employee Types: head_office, franchise
# =====================================================================
EMP_PIPELINE_STAGES = [
    "new", "qualified", "hr", "manager", "selected",
    "three_months", "joined", "hold", "rejected",
]
EMP_LINEAR_HO = ["new", "qualified", "hr", "manager", "selected", "three_months", "joined"]
EMP_LINEAR_FRANCHISE = ["new", "qualified", "hr", "selected", "three_months", "joined"]
EMP_PARALLEL = ["hold", "rejected"]
EMP_TYPES = ["head_office", "franchise"]


class EmployeeConvert(BaseModel):
    joining_date: str
    role: str
    branch_id: Optional[str] = None
    department: Optional[str] = None
    category: Optional[str] = None  # legacy: "branch" | "head_office"
    employment_type: Optional[str] = None
    salary: Optional[float] = None
    reporting_manager_id: Optional[str] = None
    employee_code: Optional[str] = None


class EmployeeCreate(BaseModel):
    name: str
    phone: Optional[str] = ""
    email: Optional[str] = ""
    employee_type: str  # "head_office" or "franchise"
    current_stage: Optional[str] = "new"
    role: Optional[str] = None  # designation name
    department: Optional[str] = None
    branch_id: Optional[str] = None
    location_city: Optional[str] = None
    location_area: Optional[str] = None
    joining_date: Optional[str] = None
    salary: Optional[float] = None
    employee_code: Optional[str] = None  # manual override
    reporting_manager_id: Optional[str] = None


class EmployeeUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    employee_type: Optional[str] = None
    role: Optional[str] = None
    department: Optional[str] = None
    branch_id: Optional[str] = None
    location_city: Optional[str] = None
    location_area: Optional[str] = None
    joining_date: Optional[str] = None
    salary: Optional[float] = None
    reporting_manager_id: Optional[str] = None


class EmployeeStageTransition(BaseModel):
    to_stage: str
    hold_reason: Optional[str] = None
    rejection_reason: Optional[str] = None


class EmployeeExit(BaseModel):
    exit_date: str
    exit_reason: str
    exit_type: str = "resigned"
    remarks: Optional[str] = ""
    auto_create_job: bool = True


async def _next_employee_code() -> str:
    """Return next EMPxxxx code (EMP0001, EMP0002, ...). Uses scan over employee_code."""
    cursor = db.employees.find({"employee_code": {"$regex": r"^EMP\d+$"}}, {"_id": 0, "employee_code": 1})
    max_num = 0
    async for doc in cursor:
        code = doc.get("employee_code") or ""
        try:
            n = int(code[3:])
            if n > max_num:
                max_num = n
        except (ValueError, IndexError):
            pass
    return f"EMP{max_num + 1:04d}"


def _derive_employee_type(lead: dict = None, branch_id: str = None, is_tech: bool = False) -> str:
    if is_tech or branch_id:
        return "franchise"
    if lead and lead.get("is_technician"):
        return "franchise"
    return "head_office"


def _derive_segmentation(lead: dict, data: EmployeeConvert) -> tuple[str, str]:
    category = data.category
    emp_type = data.employment_type
    is_tech = bool(lead.get("is_technician"))
    has_branch = bool(data.branch_id)
    if not category:
        category = "branch" if (is_tech or has_branch) else "head_office"
    if not emp_type:
        if category == "branch":
            emp_type = "technician" if is_tech else "management"
        else:
            emp_type = "management" if (data.role or "").lower() in {
                "ceo", "super admin", "hr", "marketing manager", "operations manager",
                "franchise development manager", "sales manager", "accounts manager",
            } else "mid_level"
    return category, emp_type


# ---------------- LIST / SUMMARY ----------------

@router.get("")
async def list_employees(
    employee_type: Optional[str] = None,  # "head_office" | "franchise"
    current_stage: Optional[str] = None,
    category: Optional[str] = None,  # legacy
    employment_type: Optional[str] = None,
    status: Optional[str] = None,
    branch_id: Optional[str] = None,
    search: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """List employees. Filter by employee_type/current_stage/status/branch_id/search."""
    query = {}
    if status:
        query["status"] = status
    if employee_type:
        query["employee_type"] = employee_type
    if current_stage:
        query["current_stage"] = current_stage
    if category:
        query["category"] = category
    if employment_type:
        query["employment_type"] = employment_type
    if branch_id:
        query["branch_id"] = branch_id
    if search:
        regex = {"$regex": search, "$options": "i"}
        query["$or"] = [
            {"name": regex}, {"role": regex}, {"location_city": regex},
            {"department": regex}, {"phone": regex}, {"email": regex},
            {"employee_code": regex},
        ]
    employees = await db.employees.find(query, {"_id": 0}).sort("created_at", -1).to_list(5000)
    return employees


@router.get("/exited")
async def list_exited(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in CEO_HR_ROLES:
        raise HTTPException(status_code=403, detail="Not authorized to view exited employees")
    employees = await db.employees.find({"status": "left"}, {"_id": 0}).sort("exit_date", -1).to_list(2000)
    return employees


@router.get("/pipeline-stats")
async def pipeline_stats(
    employee_type: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """Stage-wise counts + HO/Franchise/Total summary for dashboard counters."""
    match = {}
    if employee_type:
        match["employee_type"] = employee_type
    pipeline = [
        {"$match": match} if match else {"$match": {}},
        {"$group": {"_id": "$current_stage", "count": {"$sum": 1}}},
    ]
    rows = await db.employees.aggregate(pipeline).to_list(100)
    stage_counts = {s: 0 for s in EMP_PIPELINE_STAGES}
    for r in rows:
        if r["_id"] in stage_counts:
            stage_counts[r["_id"]] = r["count"]

    total = await db.employees.count_documents({})
    ho_total = await db.employees.count_documents({"employee_type": "head_office"})
    franchise_total = await db.employees.count_documents({"employee_type": "franchise"})
    joined = await db.employees.count_documents({"current_stage": "joined"})
    hold = await db.employees.count_documents({"current_stage": "hold"})
    rejected = await db.employees.count_documents({"current_stage": "rejected"})

    return {
        "stage_counts": stage_counts,
        "summary": {
            "total": total,
            "head_office": ho_total,
            "franchise": franchise_total,
            "joined": joined,
            "hold": hold,
            "rejected": rejected,
        },
    }


@router.get("/segments/summary")
async def segments_summary(current_user: dict = Depends(get_current_user)):
    """Legacy counts per segment."""
    pipeline = [
        {"$match": {"status": {"$ne": "left"}}},
        {"$group": {"_id": {"category": "$category", "employment_type": "$employment_type"}, "count": {"$sum": 1}}},
    ]
    rows = await db.employees.aggregate(pipeline).to_list(100)
    summary = {
        "branch": {"technician": 0, "management": 0},
        "head_office": {"mid_level": 0, "management": 0},
        "total_active": 0,
    }
    for r in rows:
        cat = r["_id"].get("category")
        et = r["_id"].get("employment_type")
        if cat in summary and et in summary[cat]:
            summary[cat][et] = r["count"]
            summary["total_active"] += r["count"]
    summary["total_exited"] = await db.employees.count_documents({"status": "left"})
    return summary


@router.get("/excel/template")
async def excel_template(current_user: dict = Depends(require_role("super", "manager"))):
    """Download a blank Excel template with valid columns + example rows."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    headers = [
        "Name", "Phone", "Email", "Employee Type", "Stage", "Designation",
        "Department", "Branch ID", "City", "Area", "Joining Date (YYYY-MM-DD)",
        "Salary", "Employee Code (optional)",
    ]
    ws.append(headers)
    # Examples
    ws.append([
        "John Doe", "9876543210", "john@example.com", "head_office", "joined",
        "Marketing Coordinator", "Marketing", "", "Mumbai", "Andheri",
        "2025-01-15", 35000, "",
    ])
    ws.append([
        "Jane Smith", "9876543211", "jane@example.com", "franchise", "selected",
        "Technician", "Operations", "", "Pune", "Kothrud",
        "2025-02-01", 25000, "EMP0099",
    ])
    # Notes sheet
    notes = wb.create_sheet("Instructions")
    notes.append(["Field", "Allowed Values"])
    notes.append(["Employee Type", "head_office, franchise"])
    notes.append(["Stage", ", ".join(EMP_PIPELINE_STAGES)])
    notes.append(["Joining Date", "Format YYYY-MM-DD"])
    notes.append(["Employee Code", "Leave blank to auto-generate as EMPxxxx"])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=employee_template.xlsx"},
    )


@router.get("/excel/export")
async def excel_export(
    employee_type: Optional[str] = None,
    current_stage: Optional[str] = None,
    current_user: dict = Depends(require_role("super", "manager")),
):
    """Export all employees (filtered) to an Excel sheet."""
    from openpyxl import Workbook
    query = {}
    if employee_type:
        query["employee_type"] = employee_type
    if current_stage:
        query["current_stage"] = current_stage
    employees = await db.employees.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    headers = [
        "Employee Code", "Name", "Phone", "Email", "Employee Type", "Stage",
        "Designation", "Department", "Branch ID", "City", "Area",
        "Joining Date", "Salary", "Status", "Created At",
    ]
    ws.append(headers)
    for e in employees:
        ws.append([
            e.get("employee_code", ""),
            e.get("name", ""),
            e.get("phone", ""),
            e.get("email", ""),
            e.get("employee_type", ""),
            e.get("current_stage", ""),
            e.get("role", ""),
            e.get("department", ""),
            e.get("branch_id", ""),
            e.get("location_city", ""),
            e.get("location_area", ""),
            e.get("joining_date", ""),
            e.get("salary", ""),
            e.get("status", ""),
            e.get("created_at", ""),
        ])
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"employees_export_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/excel/import")
async def excel_import(
    file: UploadFile = File(...),
    current_user: dict = Depends(require_role("super", "manager")),
):
    """Import employees from Excel. Returns created/skipped counts + errors per row."""
    from openpyxl import load_workbook
    if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx file")
    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {e}")
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Sheet is empty")
    header_map = {str(h).strip().lower(): i for i, h in enumerate(rows[0]) if h is not None}

    def col(row, *names):
        for n in names:
            idx = header_map.get(n.lower())
            if idx is not None and idx < len(row):
                v = row[idx]
                if v is None:
                    continue
                return str(v).strip() if not isinstance(v, (int, float)) else v
        return None

    created = 0
    skipped = 0
    errors: List[dict] = []
    now = datetime.now(timezone.utc).isoformat()

    for i, row in enumerate(rows[1:], start=2):
        if not any(v not in (None, "") for v in row):
            continue
        name = col(row, "Name")
        if not name:
            errors.append({"row": i, "error": "Name is required"})
            skipped += 1
            continue
        emp_type = (col(row, "Employee Type") or "head_office").lower().replace(" ", "_")
        if emp_type not in EMP_TYPES:
            errors.append({"row": i, "error": f"Invalid Employee Type '{emp_type}'"})
            skipped += 1
            continue
        stage = (col(row, "Stage") or "new").lower().replace(" ", "_")
        if stage == "3_months":
            stage = "three_months"
        if stage not in EMP_PIPELINE_STAGES:
            errors.append({"row": i, "error": f"Invalid Stage '{stage}'"})
            skipped += 1
            continue
        manual_code = col(row, "Employee Code (optional)", "Employee Code")
        if manual_code:
            manual_code = str(manual_code).strip()
            if manual_code:
                clash = await db.employees.find_one({"employee_code": manual_code})
                if clash:
                    errors.append({"row": i, "error": f"Duplicate Employee Code '{manual_code}'"})
                    skipped += 1
                    continue
                employee_code = manual_code
            else:
                employee_code = await _next_employee_code()
        else:
            employee_code = await _next_employee_code()

        joining_date = col(row, "Joining Date (YYYY-MM-DD)", "Joining Date")
        if joining_date and not isinstance(joining_date, str):
            try:
                joining_date = joining_date.strftime("%Y-%m-%d")
            except Exception:
                joining_date = str(joining_date)

        salary_val = col(row, "Salary")
        try:
            salary = float(salary_val) if salary_val not in (None, "") else None
        except Exception:
            salary = None

        emp = {
            "id": str(uuid.uuid4()),
            "employee_code": employee_code,
            "name": str(name),
            "phone": str(col(row, "Phone") or ""),
            "email": str(col(row, "Email") or ""),
            "employee_type": emp_type,
            "current_stage": stage,
            "role": str(col(row, "Designation") or ""),
            "department": str(col(row, "Department") or ""),
            "branch_id": str(col(row, "Branch ID") or "") or None,
            "location_city": str(col(row, "City") or ""),
            "location_area": str(col(row, "Area") or ""),
            "joining_date": joining_date,
            "salary": salary,
            "status": "left" if stage == "rejected" else "active",
            "source": "excel_import",
            "category": "branch" if emp_type == "franchise" else "head_office",
            "created_by": current_user["id"],
            "created_at": now,
            "updated_at": now,
        }
        await db.employees.insert_one(emp)
        created += 1

    await log_audit(
        current_user["id"], current_user["name"], "excel_import", "employee", None,
        {"created": created, "skipped": skipped, "errors": len(errors)},
    )
    return {"created": created, "skipped": skipped, "errors": errors[:50]}


# ---------------- DETAIL ----------------

@router.get("/{employee_id}")
async def get_employee(employee_id: str, current_user: dict = Depends(get_current_user)):
    e = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not e:
        raise HTTPException(status_code=404, detail="Employee not found")
    return e


# ---------------- MANUAL CREATE / UPDATE ----------------

@router.post("")
async def create_employee(
    data: EmployeeCreate,
    current_user: dict = Depends(require_role("super", "manager")),
):
    if data.employee_type not in EMP_TYPES:
        raise HTTPException(status_code=400, detail=f"Invalid employee_type. Allowed: {EMP_TYPES}")
    stage = data.current_stage or "new"
    if stage not in EMP_PIPELINE_STAGES:
        raise HTTPException(status_code=400, detail=f"Invalid stage '{stage}'")
    # Employee code
    code = (data.employee_code or "").strip()
    if code:
        clash = await db.employees.find_one({"employee_code": code})
        if clash:
            raise HTTPException(status_code=400, detail=f"Employee Code '{code}' already exists")
    else:
        code = await _next_employee_code()

    now = datetime.now(timezone.utc).isoformat()
    emp = {
        "id": str(uuid.uuid4()),
        "employee_code": code,
        "name": data.name,
        "phone": data.phone or "",
        "email": data.email or "",
        "employee_type": data.employee_type,
        "current_stage": stage,
        "role": data.role,
        "department": data.department,
        "branch_id": data.branch_id,
        "location_city": data.location_city,
        "location_area": data.location_area,
        "joining_date": data.joining_date,
        "salary": data.salary,
        "reporting_manager_id": data.reporting_manager_id,
        "status": "left" if stage == "rejected" else "active",
        "source": "manual",
        "category": "branch" if data.employee_type == "franchise" else "head_office",
        "created_by": current_user["id"],
        "created_at": now,
        "updated_at": now,
    }
    await db.employees.insert_one(emp)
    emp.pop("_id", None)
    # Stage log
    await db.employee_stage_logs.insert_one({
        "id": str(uuid.uuid4()),
        "employee_id": emp["id"],
        "from_stage": None,
        "to_stage": stage,
        "changed_by": current_user["id"],
        "changed_by_name": current_user["name"],
        "timestamp": now,
    })
    await log_audit(current_user["id"], current_user["name"], "create", "employee", emp["id"], {"name": data.name})
    return emp


@router.put("/{employee_id}")
async def update_employee(
    employee_id: str,
    data: EmployeeUpdate,
    current_user: dict = Depends(require_role("super", "manager")),
):
    payload = {k: v for k, v in data.model_dump(exclude_unset=True).items() if v is not None}
    if not payload:
        raise HTTPException(status_code=400, detail="Nothing to update")
    if "employee_type" in payload and payload["employee_type"] not in EMP_TYPES:
        raise HTTPException(status_code=400, detail="Invalid employee_type")
    payload["updated_at"] = datetime.now(timezone.utc).isoformat()
    res = await db.employees.update_one({"id": employee_id}, {"$set": payload})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Employee not found")
    await log_audit(current_user["id"], current_user["name"], "update", "employee", employee_id, payload)
    return await db.employees.find_one({"id": employee_id}, {"_id": 0})


@router.post("/{employee_id}/transition")
async def transition_employee_stage(
    employee_id: str,
    data: EmployeeStageTransition,
    current_user: dict = Depends(require_role("super", "manager")),
):
    emp = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    to_stage = data.to_stage
    if to_stage not in EMP_PIPELINE_STAGES:
        raise HTTPException(status_code=400, detail=f"Invalid stage '{to_stage}'")
    if to_stage == emp.get("current_stage"):
        raise HTTPException(status_code=400, detail="Already in this stage")
    if to_stage == "hold" and not (data.hold_reason or "").strip():
        raise HTTPException(status_code=400, detail="Hold reason is required")
    if to_stage == "rejected" and not (data.rejection_reason or "").strip():
        raise HTTPException(status_code=400, detail="Rejection reason is required")

    now = datetime.now(timezone.utc).isoformat()
    update = {
        "current_stage": to_stage,
        "updated_at": now,
    }
    if to_stage == "hold":
        update["hold_reason"] = data.hold_reason.strip()
        update["hold_at"] = now
    if to_stage == "rejected":
        update["rejection_reason"] = data.rejection_reason.strip()
        update["rejected_at"] = now
        update["status"] = "left"
    if to_stage == "joined":
        update["status"] = "active"
        if not emp.get("joining_date"):
            update["joining_date"] = now[:10]

    await db.employees.update_one({"id": employee_id}, {"$set": update})
    await db.employee_stage_logs.insert_one({
        "id": str(uuid.uuid4()),
        "employee_id": employee_id,
        "from_stage": emp.get("current_stage"),
        "to_stage": to_stage,
        "changed_by": current_user["id"],
        "changed_by_name": current_user["name"],
        "hold_reason": data.hold_reason,
        "rejection_reason": data.rejection_reason,
        "timestamp": now,
    })
    await log_audit(
        current_user["id"], current_user["name"], "employee_stage_transition", "employee", employee_id,
        {"from": emp.get("current_stage"), "to": to_stage},
    )
    return await db.employees.find_one({"id": employee_id}, {"_id": 0})


@router.get("/{employee_id}/history")
async def employee_history(employee_id: str, current_user: dict = Depends(get_current_user)):
    logs = await db.employee_stage_logs.find({"employee_id": employee_id}, {"_id": 0}).sort("timestamp", -1).to_list(200)
    return logs


class EmployeeNoteCreate(BaseModel):
    text: str


@router.get("/{employee_id}/notes")
async def list_employee_notes(employee_id: str, current_user: dict = Depends(get_current_user)):
    notes = await db.employee_notes.find({"employee_id": employee_id}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return notes


@router.post("/{employee_id}/notes")
async def add_employee_note(
    employee_id: str,
    data: EmployeeNoteCreate,
    current_user: dict = Depends(get_current_user),
):
    if not (data.text or "").strip():
        raise HTTPException(status_code=400, detail="Note text is required")
    emp = await db.employees.find_one({"id": employee_id}, {"_id": 0, "id": 1})
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    note = {
        "id": str(uuid.uuid4()),
        "employee_id": employee_id,
        "text": data.text.strip(),
        "created_by": current_user["id"],
        "created_by_name": current_user["name"],
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.employee_notes.insert_one(note)
    note.pop("_id", None)
    return note


@router.delete("/{employee_id}/notes/{note_id}")
async def delete_employee_note(
    employee_id: str,
    note_id: str,
    current_user: dict = Depends(require_role("super", "manager")),
):
    res = await db.employee_notes.delete_one({"id": note_id, "employee_id": employee_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Note not found")
    return {"success": True}


# ---------------- LEGACY: convert lead → employee ----------------

@router.post("/convert/{lead_id}")
async def convert_lead_to_employee(
    lead_id: str,
    data: EmployeeConvert,
    current_user: dict = Depends(require_role("super", "manager")),
):
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    allowed_stages = {"selected", "joined", "move_ahead", "interview_cleared", "three_months"}
    if lead["current_stage"] not in allowed_stages:
        raise HTTPException(status_code=400, detail="Lead must be in 'Selected' or 'Joined' stage to convert")
    existing = await db.employees.find_one({"lead_id": lead_id})
    if existing:
        raise HTTPException(status_code=400, detail="Lead already converted to employee")

    category, emp_type = _derive_segmentation(lead, data)
    new_employee_type = "franchise" if category == "branch" else "head_office"
    code = (data.employee_code or "").strip() or await _next_employee_code()
    if data.employee_code:
        clash = await db.employees.find_one({"employee_code": code})
        if clash:
            raise HTTPException(status_code=400, detail=f"Employee Code '{code}' already exists")

    emp_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    employee = {
        "id": emp_id,
        "lead_id": lead_id,
        "employee_code": code,
        "name": lead["name"],
        "phone": lead.get("phone", ""),
        "email": lead.get("email", ""),
        "joining_date": data.joining_date,
        "role": data.role,
        "branch_id": data.branch_id,
        "department": data.department,
        "employee_type": new_employee_type,
        "current_stage": "joined",
        "category": category,
        "employment_type": emp_type,
        "salary": data.salary,
        "reporting_manager_id": data.reporting_manager_id,
        "location_city": lead.get("location_city"),
        "location_area": lead.get("location_area"),
        "status": "active",
        "source": "lead_convert",
        "created_by": current_user["id"],
        "created_at": now,
        "updated_at": now,
    }
    await db.employees.insert_one(employee)
    employee.pop("_id", None)
    await db.employee_stage_logs.insert_one({
        "id": str(uuid.uuid4()),
        "employee_id": emp_id,
        "from_stage": None,
        "to_stage": "joined",
        "changed_by": current_user["id"],
        "changed_by_name": current_user["name"],
        "timestamp": now,
    })
    if lead["current_stage"] != "joined":
        await db.leads.update_one({"id": lead_id}, {"$set": {"current_stage": "joined", "updated_at": now}})
        await db.lead_stage_logs.insert_one({
            "id": str(uuid.uuid4()),
            "lead_id": lead_id,
            "from_stage": lead["current_stage"],
            "to_stage": "joined",
            "changed_by": current_user["id"],
            "changed_by_name": current_user["name"],
            "form_data": {"joining_date": data.joining_date},
            "timestamp": now,
        })
    await log_audit(
        current_user["id"], current_user["name"], "convert_employee", "employee", emp_id,
        {"lead_id": lead_id, "category": category, "employment_type": emp_type},
    )
    return employee


@router.post("/{employee_id}/exit")
async def mark_employee_exit(
    employee_id: str,
    data: EmployeeExit,
    current_user: dict = Depends(require_role("super")),
):
    emp = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    if emp.get("status") == "left":
        raise HTTPException(status_code=400, detail="Employee already marked as left")

    now = datetime.now(timezone.utc).isoformat()
    await db.employees.update_one(
        {"id": employee_id},
        {"$set": {
            "status": "left",
            "exit_date": data.exit_date,
            "exit_reason": data.exit_reason,
            "exit_type": data.exit_type,
            "exit_remarks": data.remarks,
            "exited_by": current_user["id"],
            "updated_at": now,
        }},
    )
    await log_audit(current_user["id"], current_user["name"], "employee_exit", "employee", employee_id,
                    {"exit_reason": data.exit_reason, "exit_type": data.exit_type})

    auto_job_id = None
    if data.auto_create_job:
        try:
            from routes.notifications import create_notification
            branch = await db.branches.find_one({"id": emp.get("branch_id")}, {"_id": 0}) if emp.get("branch_id") else None
            location = f"{branch['city']}, {branch['area']}" if branch else (emp.get("location_city") or "Head Office")
            job_id = str(uuid.uuid4())
            is_tech = emp.get("employment_type") == "technician" or emp.get("employee_type") == "franchise"
            new_job = {
                "id": job_id,
                "role": emp.get("role") or "Position",
                "type": "branch" if is_tech else "head_office",
                "department": emp.get("department"),
                "branch_id": emp.get("branch_id"),
                "location": location,
                "status": "open",
                "deadline": None,
                "auto_created_from_exit": True,
                "exit_employee_id": employee_id,
                "exit_employee_name": emp.get("name"),
                "created_by": current_user["id"],
                "created_by_name": current_user["name"],
                "created_at": now,
                "updated_at": now,
            }
            await db.jobs.insert_one(new_job)
            auto_job_id = job_id
            hr_users = await db.users.find(
                {"role": {"$in": ["Sr HR", "Jr HR"]}, "is_active": True},
                {"_id": 0, "id": 1, "name": 1},
            ).to_list(50)
            for u in hr_users:
                await create_notification(u["id"], f"New Opening — {emp.get('role')}",
                                          f"{emp.get('name')} exited from {location}. Auto job posting created.")
        except Exception as ex:
            import logging
            logging.getLogger(__name__).warning(f"Auto job creation failed: {ex}")

    try:
        from services.whatsapp import send_exit_feedback
        from routes.feedback import create_feedback_token
        token = await create_feedback_token("exit", employee_id, "employee",
                                            {"exit_reason": data.exit_reason, "exit_type": data.exit_type})
        await send_exit_feedback(emp, data.exit_reason, token)
    except Exception as ex:
        import logging
        logging.getLogger(__name__).warning(f"WhatsApp exit dispatch failed: {ex}")

    return {"success": True, "employee_id": employee_id, "auto_job_id": auto_job_id}


@router.delete("/{employee_id}")
async def delete_employee(
    employee_id: str,
    current_user: dict = Depends(require_role("super")),
):
    """Hard delete an employee (Super/CEO only). Also clears stage logs."""
    emp = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    await db.employees.delete_one({"id": employee_id})
    await db.employee_stage_logs.delete_many({"employee_id": employee_id})
    await log_audit(current_user["id"], current_user["name"], "delete", "employee", employee_id, {"name": emp.get("name")})
    return {"success": True, "deleted_employee_id": employee_id}


# ---------------- MIGRATION (idempotent) ----------------

async def migrate_employees_to_pipeline():
    """Backfill: ensures every existing employee has
       - employee_type (head_office / franchise) — derived from category if missing
       - current_stage (defaults to 'joined')
       - status ('active' if not exited, 'left' otherwise)
       - employee_code (auto EMPxxxx if missing)
    Safe to re-run.
    """
    import logging
    logger = logging.getLogger(__name__)
    cursor = db.employees.find({}, {"_id": 0})
    employee_codes_in_use: set[str] = set()
    next_seq = 0
    # First pass: collect existing codes
    async for e in db.employees.find({"employee_code": {"$regex": r"^EMP\d+$"}}, {"_id": 0, "employee_code": 1}):
        code = e.get("employee_code") or ""
        employee_codes_in_use.add(code)
        try:
            n = int(code[3:])
            if n > next_seq:
                next_seq = n
        except (ValueError, IndexError):
            pass

    migrated = 0
    async for e in cursor:
        update = {}
        # employee_type
        if not e.get("employee_type"):
            cat = e.get("category")
            if cat == "branch":
                update["employee_type"] = "franchise"
            elif cat == "head_office":
                update["employee_type"] = "head_office"
            elif e.get("branch_id"):
                update["employee_type"] = "franchise"
            else:
                update["employee_type"] = "head_office"
        # current_stage
        if not e.get("current_stage"):
            update["current_stage"] = "joined"
        # status
        if not e.get("status"):
            update["status"] = "active"
        # employee_code
        if not e.get("employee_code"):
            next_seq += 1
            new_code = f"EMP{next_seq:04d}"
            while new_code in employee_codes_in_use:
                next_seq += 1
                new_code = f"EMP{next_seq:04d}"
            employee_codes_in_use.add(new_code)
            update["employee_code"] = new_code

        if update:
            update["updated_at"] = datetime.now(timezone.utc).isoformat()
            await db.employees.update_one({"id": e["id"]}, {"$set": update})
            migrated += 1
    if migrated:
        logger.info(f"Employee pipeline migration: updated {migrated} record(s)")
    return migrated
